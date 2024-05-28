import openpyxl
import pandas as pd
from openpyxl.styles import Border, Side, Alignment, PatternFill

# Загрузка файла Excel в DataFrame
xls = pd.ExcelFile(r'C:\Users\Сергей\Desktop\full\1\forD\updated_сектор.xlsx')
df = pd.DataFrame()

# Объединение всех листов в один DataFrame
for sheet_name in xls.sheet_names:
    df_temp = xls.parse(sheet_name)
    df = pd.concat([df, df_temp], ignore_index=True)
    df = df.fillna(0)
    df['Осталось продать'] = df.iloc[:, 3] - df.iloc[:, 2]
    df = df.sort_values(by='Осталось продать', ascending=True)

# Удаление столбца "Планы"
df = df.drop(columns=['Планы'])

# Вычисление итоговых значений
df_length = len(df)
df['Итог план'] = df.iloc[:, 2].sum()
df['Итого осталось продать'] = df.iloc[:, 3].sum() - df['Итог план']
df['Итого продано'] = df['Итог план'] + df['Итого осталось продать']

# Удаление всех дубликатов из DataFrame
df = df.drop_duplicates()

# Фильтрация клиентов с отрицательным значением и с нулевым значением
df_negative = df[df['Осталось продать'] < 0]
df_zero = df[df['Осталось продать'] == 0]

# Запись данных в новый Excel-файл
with pd.ExcelWriter(r'C:\Users\Сергей\Desktop\full\1\сектор_1.xlsx', engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Данные')
    df_negative.to_excel(writer, index=False, sheet_name='Отрицательные')
    workbook = writer.book
    worksheet_data = writer.sheets['Данные']
    worksheet_negative = writer.sheets['Отрицательные']

    # Применение одинакового стиля к ячейкам на обоих листах
    for worksheet in [worksheet_data, worksheet_negative]:
        for row in range(1, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                if cell.value != 0:
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Автоподбор ширины столбцов
    for col in range(1, len(df.columns) + 1):
        max_length = 0
        for row in worksheet_data.iter_rows(min_row=1, max_row=len(df) + 1, min_col=col, max_col=col):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        worksheet_data.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_length + 2
        worksheet_negative.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_length + 2

print('Выполнено успешно 1 сектор')
